"""Excel intake parser.

Parses an uploaded .xlsx file into a list of `ChangeLineInput` records the
intake validator can score.

Required headers: `category`, `action`, `pipeline_layer`, `entity` (case-
insensitive). Optional headers: `target_attribute`, `target_dataset`,
`target_table`, `target_column`, `target_data_type`, `target_nullable`,
`source_system`, `source_dataset`, `source_table`, `source_column`,
`transformation_logic`, `business_definition`, `rationale`,
`impact_notes`.

Rows missing any required field are skipped silently — Excel templates
tend to have placeholder rows we don't want to surface as errors. Rows
with unknown `category` or `action` values are skipped the same way; the
preview-then-submit flow gives users a chance to fix them upstream.
"""
from __future__ import annotations

from io import BytesIO

import structlog
from openpyxl import load_workbook

from app.domain.requests import (
    ChangeAction,
    ChangeCategory,
    ChangeLineInput,
    PipelineLayer,
)

logger = structlog.get_logger(__name__)

_SUPPORTED_COLUMNS = {
    "category",
    "action",
    "pipeline_layer",
    "entity",
    "target_attribute",
    "target_dataset",
    "target_table",
    "target_column",
    "target_data_type",
    "target_nullable",
    "source_system",
    "source_dataset",
    "source_table",
    "source_column",
    "transformation_logic",
    "business_definition",
    "rationale",
    "impact_notes",
}

_REQUIRED = ("category", "action", "pipeline_layer", "entity")


class ExcelParseError(ValueError):
    """Raised when the workbook is missing required headers or unreadable."""


def parse_excel(content: bytes) -> list[ChangeLineInput]:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Cannot read uploaded file as .xlsx: {exc}") from exc

    sheet = wb.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise ExcelParseError("Workbook has no rows")

    headers = [
        (str(h).strip().lower() if h is not None else "") for h in header_row
    ]
    missing = [h for h in _REQUIRED if h not in headers]
    if missing:
        raise ExcelParseError(
            f"Header row must include required columns: {', '.join(missing)}"
        )

    valid_categories = {c.value for c in ChangeCategory}
    valid_actions = {a.value for a in ChangeAction}
    valid_layers = {p.value for p in PipelineLayer}

    items: list[ChangeLineInput] = []

    for raw in rows:
        record: dict[str, str | None] = {}
        for idx, header in enumerate(headers):
            if header not in _SUPPORTED_COLUMNS:
                continue
            value = raw[idx] if idx < len(raw) else None
            if value is None or (isinstance(value, str) and not value.strip()):
                record[header] = None
            else:
                record[header] = str(value).strip()

        if any(not record.get(h) for h in _REQUIRED):
            continue
        if (
            record["category"] not in valid_categories
            or record["action"] not in valid_actions
            or record["pipeline_layer"] not in valid_layers
        ):
            continue

        nullable_raw = record.get("target_nullable")
        target_nullable: bool | None = None
        if nullable_raw is not None:
            target_nullable = nullable_raw.lower() in ("true", "1", "yes", "y")

        items.append(
            ChangeLineInput(
                category=ChangeCategory(record["category"]),
                action=ChangeAction(record["action"]),
                pipeline_layer=PipelineLayer(record["pipeline_layer"]),
                entity=record["entity"] or "",
                target_attribute=record.get("target_attribute"),
                target_dataset=record.get("target_dataset"),
                target_table=record.get("target_table"),
                target_column=record.get("target_column"),
                target_data_type=record.get("target_data_type"),
                target_nullable=target_nullable,
                source_system=record.get("source_system"),
                source_dataset=record.get("source_dataset"),
                source_table=record.get("source_table"),
                source_column=record.get("source_column"),
                transformation_logic=record.get("transformation_logic"),
                business_definition=record.get("business_definition"),
                rationale=record.get("rationale"),
                impact_notes=record.get("impact_notes"),
            )
        )

    logger.info("excel_parsed", row_count=len(items))
    return items
